"""
extract.py  --  Read DMA data from Excel and populate SQLite database.

Usage:
    python scripts/extract.py data/dma_tatas.xlsx build/dma.db
"""

import sys
import sqlite3
import openpyxl

# Maps exact sheet name -> (sample, axis)
SHEET_MAP = {
    "aquamarine product board x-axis": ("product_board", "x"),
    "aquamarin product board yaxis":   ("product_board", "y"),
    "aquamarine product board-zaxis":  ("product_board", "z"),
    "aquamarine solder board x-axis":  ("solder_board",  "x"),
    "aquamarine solder board y axis":  ("solder_board",  "y"),
    "aquamarine solder board z-axis":  ("solder_board",  "z"),
}

REQUIRED_COLS = [
    "Sample Temperature",
    "Frequency",
    "Storage Modulus",
    "Loss Modulus",
    "Tan Delta",
]


def create_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dma_measurements (
            id               INTEGER PRIMARY KEY,
            sample           TEXT    NOT NULL,
            axis             TEXT    NOT NULL,
            temperature      REAL,
            frequency        REAL,
            storage_modulus  REAL,
            loss_modulus     REAL,
            tan_delta        REAL
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_sample_axis
        ON dma_measurements (sample, axis)
    """)
    conn.commit()


def load_sheet(ws, sample, axis):
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]

    try:
        idx = {col: header.index(col) for col in REQUIRED_COLS}
    except ValueError as e:
        print(f"  WARNING: missing column in {sample}/{axis}: {e}")
        return []

    records = []
    for row in rows:
        temp  = row[idx["Sample Temperature"]]
        freq  = row[idx["Frequency"]]
        e_p   = row[idx["Storage Modulus"]]
        e_pp  = row[idx["Loss Modulus"]]
        tand  = row[idx["Tan Delta"]]

        if temp is None:
            continue

        records.append((sample, axis, temp, freq, e_p, e_pp, tand))

    return records


def main(xlsx_path, db_path):
    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    conn = sqlite3.connect(db_path)
    create_table(conn)

    total = 0
    for sheet_name, (sample, axis) in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet not found: {sheet_name}")
            continue

        print(f"  Loading sheet '{sheet_name}' -> sample={sample}, axis={axis}")
        ws = wb[sheet_name]
        records = load_sheet(ws, sample, axis)

        conn.executemany(
            "INSERT INTO dma_measurements "
            "(sample, axis, temperature, frequency, storage_modulus, loss_modulus, tan_delta) "
            "VALUES (?,?,?,?,?,?,?)",
            records,
        )
        conn.commit()
        print(f"    Inserted {len(records)} rows")
        total += len(records)

    wb.close()
    conn.close()
    print(f"\nDone. Total rows inserted: {total}")
    print(f"Database saved to: {db_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python scripts/extract.py <xlsx_path> <db_path>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
